#!/usr/bin/env python3
import socket
import sys
import requests
import validators
import json
from openpyxl import load_workbook, Workbook
from cryptography import x509

# Please `pip install openpyxl requests validators` before running

# Global variables as settings
evoidfile = 'evoid'
blacklistf = 'blacklist'
TIMEOUT = 4
socket.setdefaulttimeout(TIMEOUT)

# Print Usage
if len(sys.argv) < 3:
	print('Usage: certfilter.py inputfile.xlsx outputfile.xlsx')
	sys.exit(1)

# Prepare set for EVSSL oids
with open(evoidfile)as f:
	evoid = set(f.read().splitlines())

with open(blacklistf)as f:
	blacklist = set(f.read().splitlines())

f.close()
column = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N')
title = ('Host', 'Subject C', 'Subject O', 'Subject OU', 'Subject CN', 'Issuer C', 'Issuer O', 'Issuer OU', 'Issuer CN',
         'notBefore', 'notAfter', 'certType', 'weight', 'taxId')


# Prevent empty extraction of cert element
def xstr(s):
	if s is None:
		return ''
	return str(s)


# Simple way to check if host accept TLS connection
def isalive(fqdn='hicloud.hinet.net', port=443, timeout=TIMEOUT):
	socket.setdefaulttimeout(timeout)
	if fqdn in blacklist:
		return False
	elif validators.url('https://' + xstr(fqdn)):
		try:
			socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((fqdn, port))
			return True
		except Exception:
			return False
	else:
		return False


# Fetch peer certificate from Internet by requests module override original
HTTPResponse = requests.packages.urllib3.response.HTTPResponse
orig_HTTPResponse__init__ = HTTPResponse.__init__


def new_HTTPResponse__init__(self, *args, **kwargs):
	orig_HTTPResponse__init__(self, *args, **kwargs)
	try:
		self.peer_certificate = self._connection.peer_certificate
	except AttributeError:
		pass


HTTPResponse.__init__ = new_HTTPResponse__init__
HTTPAdapter = requests.adapters.HTTPAdapter
orig_HTTPAdapter_build_response = HTTPAdapter.build_response


def new_HTTPAdapter_build_response(self, request, resp):
	response = orig_HTTPAdapter_build_response(self, request, resp)
	try:
		response.peer_certificate = resp.peer_certificate
	except AttributeError:
		pass
	return response


HTTPAdapter.build_response = new_HTTPAdapter_build_response
HTTPSConnection = requests.packages.urllib3.connection.HTTPSConnection
orig_HTTPSConnection_connect = HTTPSConnection.connect


def new_HTTPSConnection_connect(self):
	orig_HTTPSConnection_connect(self)
	try:
		self.peer_certificate = self.sock.connection.get_peer_certificate()
	except AttributeError:
		pass


HTTPSConnection.connect = new_HTTPSConnection_connect


# Get Business_Accounting_NO from gcis
def getbano(corporate):
	if corporate is None:
		return None
	else:
		try:
			gcisapi = 'http://data.gcis.nat.gov.tw/od/data/api/6BBA2268-1367-4B42-9CCA-BC17499EBE8C?$format=json' \
			          '&$filter=Company_Name like ' + corporate + ' and Company_Status eq 01&$skip=0&$top=1'
			# https://data.gcis.nat.gov.tw/od/demo_cond/6BBA2268-1367-4B42-9CCA-BC17499EBE8C for reference
			r = requests.get(gcisapi)
			js = json.loads(r.text)
			no = js[0]['Business_Accounting_NO']
			return no
		except Exception:
			return None


# Worksheet preparation
out_wb = Workbook()
input_file = sys.argv[1]
output_name = sys.argv[2]
in_wb = load_workbook(filename=input_file)
in_ws = in_wb[in_wb.sheetnames[0]]
out_ws = out_wb.active
out_ws.title = 'scanned_result'
for i in range(14):
	out_ws[column[i] + str(1)] = title[i]

# Worker loop
for row in range(2, in_ws.max_row + 1):
	sub_c = sub_o = sub_ou = sub_cn = sub_l = sub_st = sub_pc = iss_c = iss_o = iss_ou = iss_cn = nota = notb = bano = certtype = None
	host = in_ws['A' + str(row)].value
	weight = in_ws['B' + str(row)].value
	if isalive(xstr(host)):
		try:
			r = requests.get('https://' + host, timeout=TIMEOUT)
			sub_c = r.peer_certificate.get_subject().C
			sub_o = r.peer_certificate.get_subject().O
			sub_ou = r.peer_certificate.get_subject().OU
			sub_cn = r.peer_certificate.get_subject().CN
			sub_l = r.peer_certificate.get_subject().L
			sub_st = r.peer_certificate.get_subject().ST
			sub_pc = r.peer_certificate.get_subject().postalCode
			iss_c = r.peer_certificate.get_issuer().C
			iss_o = r.peer_certificate.get_issuer().O
			iss_ou = r.peer_certificate.get_issuer().OU
			iss_cn = r.peer_certificate.get_issuer().CN
			notb = r.peer_certificate.get_notBefore().decode()
			nota = r.peer_certificate.get_notAfter().decode()
			bano = getbano(sub_o)
			cert = r.peer_certificate.to_cryptography()
			policy = cert.extensions.get_extension_for_class(x509.CertificatePolicies)
			for i in range(len(policy.value) - 1, -1, -1):
				if policy.value[i].policy_qualifiers is None:
					certtype = \
						{'2.23.140.1.1': 'EV', '2.23.140.1.2.1': 'DV', '2.23.140.1.2.2': 'OV', '2.23.140.1.2.3': 'IV'}[
							policy.value[i].policy_identifier.dotted_string]
					break
				elif policy.value[i].policy_identifier.dotted_string in evoid:
					certtype = 'EV'
					break
				elif sub_o == sub_cn or 'Domain Control Validated' in sub_ou:
					if sub_l is None and sub_st is None and sub_pc is None:
						certtype = 'DV'
						break
				elif sub_o is None and iss_o is 'StartCom':
					certtype = 'DV'
					break

			if certtype is None:
				if sub_o is None:
					certtype = 'DV'
				else:
					certtype = 'OV'

			print(str(row) + ',' + xstr(host) + ',' + xstr(sub_c) + ',' + xstr(sub_o) + ',' + xstr(sub_ou) + ',' + xstr(
				sub_cn) + ',' + xstr(iss_c) + ',' + xstr(iss_o) + ',' + xstr(iss_ou) + ',' + xstr(iss_cn) + ',' + xstr(
				notb) + ',' + xstr(nota) + ',' + str(certtype) + ',' + str(weight) + ',' + xstr(bano))

			content = [host, sub_c, sub_o, sub_ou, sub_cn, iss_c, iss_o, iss_ou, iss_cn, notb, nota, certtype, weight,
			           bano]
			for i in range(14):
				out_ws[column[i] + str(row)] = content[i]

		except Exception:
			pass
	else:
		print(str(row) + ',' + xstr(host) + 'Not Available')
		content = [host, None, None, None, None, None, None, None, None, None, None, None, weight, None]
		for i in range(14):
			out_ws[column[i] + str(row)] = content[i]
out_wb.save(filename=output_name)
